"""Konferensiyalar katalogi (conferences_bp) — mahalliy + xalqaro.

Ikkita katalog: mahalliy (vazirlik yillik rejasi, scripts/scrape_conferences_local.py)
va xalqaro (confs.tech ochiq datasetlari, scripts/scrape_conferences_intl.py).
Statik ro'yxatdan farqi — sayt tizimlariga integratsiya:

  * bookmark (user_conference_bookmarks — dissertatsiya user_bookmarks UX'i
    aynan takrorlanadi, lekin FK conferences ga; jadval alohida, chunki
    user_bookmarks.dissertation_id NOT NULL FK va polimorf emas);
  * "Roadmap'ga qo'shish" — mavjud roadmap_conferences jadvaliga bitta bosishda
    yozadi (source_conference_id bilan; qo'lda kiritish o'zgarishsiz ishlaydi);
  * ixtisoslik obunasi — yangi konferensiyalar mos obunachilarga user_alerts +
    Telegram orqali boradi (subscriptions.notify_specialty_subscribers naqshi,
    dedup conference_notifications_log da).

Sxema lazy + idempotent (_ensure_schema) — migrations/add_conferences.sql bilan
bir xil, server birinchi so'rovda self-migrate (grants/reminders konvensiyasi).

Routes:
  GET  /konferensiyalar                — hub: 2 katta karta (jonli sonlar bilan)
  GET  /konferensiyalar/mahalliy       — mahalliy katalog
  GET  /konferensiyalar/xalqaro        — xalqaro katalog
  GET  /konferensiya/<slug>            — detail sahifa
  GET  /api/v1/conferences             — filtrlanadigan JSON (facet countlar bilan)
  POST /api/v1/conferences/bookmark    — ⭐ toggle {conference_id}
  POST /api/v1/conferences/roadmap     — Roadmap'ga qo'shish {conference_id}
  POST /api/v1/conferences/dispatch-alerts — yangi konf. obunachilarga (API key,
                                         GitHub Actions scraperdan keyin chaqiradi)
  /admin/conferences*                  — admin CRUD (admin/grants naqshi)
"""
import os
import re
from datetime import date, datetime

from flask import (Blueprint, jsonify, request, render_template,
                   redirect, abort, flash)
from flask_login import login_required, current_user

from app import csrf
from data import get_connection

try:
    import psycopg2.extras as psycopg2_extras
except Exception:  # pragma: no cover — psycopg2 always present in prod
    psycopg2_extras = None

conferences_bp = Blueprint('conferences', __name__)

_schema_ready = False

PER_PAGE = 12
SIMILAR_LIMIT = 4

SCOPES = ('local', 'international')
FORMATS = {'onsite': 'Anʼanaviy', 'online': 'Onlayn', 'hybrid': 'Gibrid'}
EVENT_TYPES_LOCAL = ('Anjuman', 'Forum', 'Kongress', 'Seminar', 'Simpozium',
                     'Konferensiya')
MONTHS_UZ = ['Yanvar', 'Fevral', 'Mart', 'Aprel', 'May', 'Iyun', 'Iyul',
             'Avgust', 'Sentabr', 'Oktabr', 'Noyabr', 'Dekabr']

# Vazirlik yo'nalishlari — conference_fields boshlang'ich seed (normalizatsiya).
CONFERENCE_FIELDS_SEED = [
    'Axborot va kommunikatsiya texnologiyalari',
    'Biologiya va biotexnologiya',
    "Fan va ta'lim",
    'Geologiya-mineralogiya, seysmologiya va zilzilaga chidamliligi, '
    'qurilish va arxitektura',
    'Ijtimoiy-gumanitar fanlar',
    'Iqtisodiyot, davlat va jamiyat qurilishi',
    'Neft va gaz, energiya, energiya va resurslarni tejash',
    "Qishloq va suv xo'jaligi",
    'Sanoat, ishlab chiqarish, transport va logistika',
    "San'at va madaniyat",
    'Tibbiyot, farmakologiya va sport',
]

# deadline eslatma oynalari (kun)
DEADLINE_BUCKETS = (14, 7, 1)


def _field_slug(name):
    from institutions import transliterate
    s = transliterate((name or '').lower())
    for ch in "'ʻʼ‘’`":
        s = s.replace(ch, '')
    s = re.sub(r'[^a-z0-9]+', '-', s).strip('-')[:280].strip('-')
    return s or 'soha'

# Obuna moslashuvi: ixtisoslik_nomi tokenlari konf. nomi/sohasida qidiriladi.
# Umumiy so'zlar chiqarib tashlanadi — mo'rt moslikni majburlamaymiz (spec).
_MATCH_STOP = {'fanlari', 'fanlar', 'boyicha', "bo'yicha", 'sohasi', 'hamda',
               'умумий', 'фанлари', 'фанлар', 'соҳаси'}
_TOKEN_RE = re.compile(r"[^\W\d_]{6,}", re.UNICODE)


# ── helpers ──────────────────────────────────────────────────────────────────

def _days_until(d):
    """Bugungacha qolgan kunlar (o'tgan bo'lsa manfiy); None — sana yo'q."""
    if not d:
        return None
    if isinstance(d, str):
        try:
            d = date.fromisoformat(d[:10])
        except ValueError:
            return None
    if isinstance(d, datetime):
        d = d.date()
    return (d - date.today()).days


def _uz_date(d):
    months = ['yanvar', 'fevral', 'mart', 'aprel', 'may', 'iyun', 'iyul',
              'avgust', 'sentabr', 'oktabr', 'noyabr', 'dekabr']
    if isinstance(d, str):
        try:
            d = date.fromisoformat(d[:10])
        except ValueError:
            return d
    if isinstance(d, datetime):
        d = d.date()
    return f'{d.day} {months[d.month - 1]} {d.year}' if d else ''


def make_slug(title, cur=None):
    """URL slug (kirill→lotin); to'qnashuvda -2, -3… (cur berilsa DB tekshiradi).
    Hech qachon faqat raqam bo'lmaydi."""
    from institutions import transliterate
    s = transliterate((title or '').lower())
    for ch in "'ʻʼ‘’`":
        s = s.replace(ch, '')
    s = re.sub(r'[^a-z0-9]+', '-', s).strip('-')[:200].strip('-')
    if not s or s.replace('-', '').isdigit():
        s = f'konf-{s}'.strip('-')
    if cur is None:
        return s
    base, n = s, 2
    while True:
        cur.execute("SELECT 1 FROM conferences WHERE title_slug = %s", (s,))
        if not cur.fetchone():
            return s
        s = f'{base}-{n}'[:250]
        n += 1


def _parse_date(v):
    if not v:
        return None
    try:
        return date.fromisoformat(str(v)[:10])
    except ValueError:
        return None


# ── schema (lazy, idempotent — mirrors migrations/add_conferences.sql) ──────

def _ensure_schema(cur):
    global _schema_ready
    if _schema_ready:
        return
    cur.execute("""
        CREATE TABLE IF NOT EXISTS conferences (
            id SERIAL PRIMARY KEY,
            title VARCHAR(600) NOT NULL,
            title_slug VARCHAR(250) UNIQUE,
            scope VARCHAR(10) NOT NULL CHECK (scope IN ('local','international')),
            organizer VARCHAR(400),
            field VARCHAR(200),
            region VARCHAR(100),
            city VARCHAR(150),
            event_type VARCHAR(50),
            start_date DATE,
            end_date DATE,
            is_multiday BOOLEAN DEFAULT FALSE,
            format VARCHAR(20),
            publisher VARCHAR(200),
            is_scopus_indexed BOOLEAN DEFAULT FALSE,
            submission_deadline DATE,
            country VARCHAR(100),
            source_url VARCHAR(500),
            source_id VARCHAR(300) UNIQUE,
            description TEXT,
            is_active BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        )""")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_conf_scope_date "
                "ON conferences(scope, start_date) WHERE is_active = TRUE")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_conf_field ON conferences(field)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_conf_region "
                "ON conferences(region) WHERE scope = 'local'")
    try:
        cur.execute("CREATE INDEX IF NOT EXISTS idx_conf_title_trgm "
                    "ON conferences USING GIN (title gin_trgm_ops)")
    except Exception:
        pass  # pg_trgm yo'q muhitda ILIKE indekssiz ishlayveradi
    cur.execute("""
        CREATE TABLE IF NOT EXISTS user_conference_bookmarks (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            conference_id INTEGER NOT NULL REFERENCES conferences(id) ON DELETE CASCADE,
            created_at TIMESTAMP DEFAULT NOW(),
            UNIQUE(user_id, conference_id)
        )""")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_conf_bm_user "
                "ON user_conference_bookmarks(user_id, created_at DESC)")
    # obuna dedup logi (specialty_notifications_log aksi)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS conference_notifications_log (
            id SERIAL PRIMARY KEY,
            subscription_id INTEGER REFERENCES specialty_subscriptions(id) ON DELETE CASCADE,
            conference_id INTEGER REFERENCES conferences(id) ON DELETE CASCADE,
            sent_via VARCHAR(20) DEFAULT 'site',
            sent_at TIMESTAMP DEFAULT NOW()
        )""")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_conf_notif_log "
                "ON conference_notifications_log(subscription_id, conference_id)")
    # Roadmap integratsiyasi: katalogdan qo'shilgan yozuvlarni belgilash.
    # Qo'lda kiritilganlar uchun NULL — mavjud xatti-harakat o'zgarmaydi.
    cur.execute("ALTER TABLE IF EXISTS roadmap_conferences "
                "ADD COLUMN IF NOT EXISTS source_conference_id INTEGER")

    # ── kengaytirish (v2): sohalar normalizatsiyasi + qo'shimcha maydonlar ──
    cur.execute("""
        CREATE TABLE IF NOT EXISTS conference_fields (
            id SERIAL PRIMARY KEY,
            name VARCHAR(300) NOT NULL,
            slug VARCHAR(300) UNIQUE
        )""")
    for _fname in CONFERENCE_FIELDS_SEED:
        cur.execute(
            "INSERT INTO conference_fields (name, slug) VALUES (%s, %s) "
            "ON CONFLICT (slug) DO NOTHING", (_fname, _field_slug(_fname)))
    # conferences'ga yangi ustunlar (mavjud ekvivalentlar qayta ishlatiladi:
    # start_date/end_date, is_scopus_indexed, title_slug, is_active, event_type)
    for _col, _typ in (
        ('acronym', 'VARCHAR(50)'), ('title_en', 'TEXT'),
        ('field_id', 'INTEGER'), ('tags', "TEXT[]"),
        ('notification_date', 'DATE'), ('registration_deadline', 'DATE'),
        ('venue', 'TEXT'), ('organizer_contact', 'TEXT'),
        ('cfp_url', 'TEXT'), ('poster_image', 'TEXT'),
        ('ccf_rank', 'VARCHAR(8)'), ('core_rank', 'VARCHAR(8)'),
        ('doi_available', 'BOOLEAN DEFAULT FALSE'),
        ('source', 'VARCHAR(50)'),
    ):
        cur.execute(f"ALTER TABLE conferences ADD COLUMN IF NOT EXISTS {_col} {_typ}")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_conf_field_id "
                "ON conferences(field_id)")
    # field_id — conference_fields'ga yumshoq havola (formal FK yo'q, repo naqshi:
    # muassasa/institution_map kabi text/soft-mode — migratsiya xavfsizroq)
    # mavjud `field` (erkin matn) → field_id backfill (nom bo'yicha moslik)
    cur.execute("""
        UPDATE conferences c SET field_id = f.id
        FROM conference_fields f
        WHERE c.field_id IS NULL AND c.field IS NOT NULL
          AND LOWER(TRIM(c.field)) = LOWER(TRIM(f.name))
    """)
    # deadline eslatma dedup logi (bookmark egasi × konf. × kun-oynasi)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS conference_deadline_log (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            conference_id INTEGER NOT NULL REFERENCES conferences(id) ON DELETE CASCADE,
            days_bucket INTEGER NOT NULL,
            sent_at TIMESTAMP DEFAULT NOW(),
            UNIQUE(user_id, conference_id, days_bucket)
        )""")
    _schema_ready = True


_CONF_COLS = ('id', 'title', 'title_slug', 'scope', 'organizer', 'field',
              'region', 'city', 'event_type', 'start_date', 'end_date',
              'is_multiday', 'format', 'publisher', 'is_scopus_indexed',
              'submission_deadline', 'country', 'source_url', 'description',
              'is_active',
              # v2 kengaytirish
              'acronym', 'title_en', 'field_id', 'tags', 'notification_date',
              'registration_deadline', 'venue', 'organizer_contact', 'cfp_url',
              'poster_image', 'ccf_rank', 'core_rank', 'doi_available')


def _card(r):
    """DB qatori (RealDict) → JSON/template item."""
    start, end = r.get('start_date'), r.get('end_date')
    sub = r.get('submission_deadline')
    days = _days_until(start)
    item = {
        'id': r['id'],
        'title': r.get('title') or '',
        'slug': r.get('title_slug') or '',
        'scope': r.get('scope'),
        'organizer': r.get('organizer') or '',
        'field': r.get('field') or '',
        'region': r.get('region') or '',
        'city': r.get('city') or '',
        'country': r.get('country') or '',
        'event_type': r.get('event_type') or '',
        'start_date': start.isoformat() if start else None,
        'end_date': end.isoformat() if end else None,
        'start_uz': _uz_date(start),
        'end_uz': _uz_date(end),
        'is_multiday': bool(r.get('is_multiday')),
        'format': r.get('format') or '',
        'format_uz': FORMATS.get(r.get('format') or '', ''),
        'publisher': r.get('publisher') or '',
        'is_scopus_indexed': bool(r.get('is_scopus_indexed')),
        'submission_deadline': sub.isoformat() if sub else None,
        'submission_uz': _uz_date(sub),
        'submission_days': _days_until(sub),
        'source_url': r.get('source_url') or '',
        'description': r.get('description') or '',
        'days_remaining': days,
        'expired': days is not None and days < 0,
        # v2 maydonlar
        'acronym': r.get('acronym') or '',
        'title_en': r.get('title_en') or '',
        'field_id': r.get('field_id'),
        'tags': list(r.get('tags') or []),
        'notification_date': (r['notification_date'].isoformat()
                              if r.get('notification_date') else None),
        'notification_uz': _uz_date(r.get('notification_date')),
        'registration_deadline': (r['registration_deadline'].isoformat()
                                  if r.get('registration_deadline') else None),
        'registration_uz': _uz_date(r.get('registration_deadline')),
        'registration_days': _days_until(r.get('registration_deadline')),
        'venue': r.get('venue') or '',
        'organizer_contact': r.get('organizer_contact') or '',
        'cfp_url': r.get('cfp_url') or '',
        'poster_image': r.get('poster_image') or '',
        'ccf_rank': (r.get('ccf_rank') or '').upper(),
        'core_rank': (r.get('core_rank') or '').upper(),
        'doi_available': bool(r.get('doi_available')),
    }
    return item


# ── ro'yxat so'rovi (filtrlar + facetlar) ────────────────────────────────────

def _build_where(scope, args, uid):
    """(where_sql, params, saved_join). Facetlar ham shu yordamchidan quriladi
    (exclude parametri bilan) — bitta haqiqat manbai."""
    where = ["is_active = TRUE", "scope = %s"]
    params = [scope]

    time_f = args.get('time') or 'upcoming'
    if time_f == 'upcoming':
        where.append("(COALESCE(end_date, start_date) >= CURRENT_DATE "
                     "OR start_date IS NULL)")
    elif time_f == 'past':
        where.append("COALESCE(end_date, start_date) < CURRENT_DATE")

    q = (args.get('search') or '').strip()
    if q:
        where.append("(title ILIKE %s OR organizer ILIKE %s OR city ILIKE %s)")
        params += [f'%{q}%'] * 3

    def _multi(name, col):
        vals = [v for v in args.getlist(name) if v]
        if vals:
            where.append(f"{col} = ANY(%s)")
            params.append(vals)

    _multi('field', 'field')
    _multi('region', 'region')
    _multi('type', 'event_type')
    _multi('publisher', 'publisher')
    _multi('format', 'format')

    month = args.get('month')
    if month and month.isdigit() and 1 <= int(month) <= 12:
        where.append("EXTRACT(MONTH FROM start_date) = %s")
        params.append(int(month))

    if args.get('scopus') in ('1', 'true'):
        where.append("is_scopus_indexed = TRUE")

    saved_join = ''
    if args.get('saved') in ('1', 'true') and uid:
        saved_join = ("JOIN user_conference_bookmarks bm "
                      "ON bm.conference_id = conferences.id AND bm.user_id = %s")
    return ' AND '.join(where), params, saved_join


_FACET_DIMS = {  # facet nomi → (query param, ustun)
    'field': ('field', 'field'),
    'region': ('region', 'region'),
    'type': ('type', 'event_type'),
    'publisher': ('publisher', 'publisher'),
    'format': ('format', 'format'),
    'month': ('month', 'EXTRACT(MONTH FROM start_date)::int'),
}


class _ArgsView:
    """request.args ko'rinishi — bitta parametr olib tashlangan (facet uchun:
    o'z filtri o'ziga ta'sir qilmasin, boshqalari qo'llansin)."""

    def __init__(self, args, drop):
        self._a, self._drop = args, drop

    def get(self, k, default=None):
        return default if k == self._drop else self._a.get(k, default)

    def getlist(self, k):
        return [] if k == self._drop else self._a.getlist(k)


@conferences_bp.route('/api/v1/conferences')
def api_conferences():
    scope = request.args.get('scope') or 'local'
    if scope not in SCOPES:
        return jsonify({'ok': False, 'error': 'scope local|international'}), 400
    uid = current_user.id if getattr(current_user, 'is_authenticated', False) else None
    try:
        page = max(1, int(request.args.get('page', 1)))
    except (TypeError, ValueError):
        page = 1

    conn = get_connection()
    try:
        cur = conn.cursor(cursor_factory=psycopg2_extras.RealDictCursor)
        _ensure_schema(cur)

        where, params, saved_join = _build_where(scope, request.args, uid)
        jp = [uid] if saved_join else []
        order = ("start_date ASC NULLS LAST, id DESC"
                 if (request.args.get('time') or 'upcoming') != 'past'
                 else "start_date DESC NULLS LAST, id DESC")

        cur.execute(f"SELECT COUNT(*) AS n FROM conferences {saved_join} "
                    f"WHERE {where}", jp + params)
        total = cur.fetchone()['n']

        cur.execute(
            f"SELECT {', '.join(_CONF_COLS)} FROM conferences {saved_join} "
            f"WHERE {where} ORDER BY {order} LIMIT %s OFFSET %s",
            jp + params + [PER_PAGE, (page - 1) * PER_PAGE])
        items = [_card(r) for r in cur.fetchall()]

        # facet countlar — har o'lchov o'z filtrisiz, qolganlari bilan
        facet_names = (('field', 'region', 'type', 'month') if scope == 'local'
                       else ('field', 'type', 'publisher', 'format', 'month'))
        facets = {}
        for name in facet_names:
            param, col = _FACET_DIMS[name]
            w, p, sj = _build_where(scope, _ArgsView(request.args, param), uid)
            jp2 = [uid] if sj else []
            cur.execute(
                f"SELECT {col} AS v, COUNT(*) AS n FROM conferences {sj} "
                f"WHERE {w} AND {col} IS NOT NULL "
                f"GROUP BY 1 ORDER BY 2 DESC, 1 LIMIT 30", jp2 + p)
            facets[name] = [{'value': str(r['v']), 'count': r['n']}
                            for r in cur.fetchall() if r['v'] not in (None, '')]

        saved_ids = []
        if uid:
            cur.execute("SELECT conference_id FROM user_conference_bookmarks "
                        "WHERE user_id = %s", (uid,))
            saved_ids = [r['conference_id'] for r in cur.fetchall()]
        conn.commit()
        pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)
        return jsonify({'ok': True, 'items': items, 'count': total,
                        'page': min(page, pages), 'pages': pages,
                        'facets': facets, 'saved_ids': saved_ids})
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'error': 'server_error'}), 500
    finally:
        conn.close()


# ── sahifalar ────────────────────────────────────────────────────────────────

def _scope_stats(cur, scope):
    cur.execute("""
        SELECT COUNT(*) AS total,
               COUNT(DISTINCT region) FILTER (WHERE region IS NOT NULL) AS regions,
               COUNT(DISTINCT country) FILTER (WHERE country IS NOT NULL) AS countries,
               COUNT(DISTINCT field) FILTER (WHERE field IS NOT NULL) AS fields,
               COUNT(*) FILTER (WHERE COALESCE(end_date, start_date) >= CURRENT_DATE) AS upcoming
        FROM conferences WHERE is_active = TRUE AND scope = %s
    """, (scope,))
    r = cur.fetchone()
    return {'total': r[0] or 0, 'regions': r[1] or 0, 'countries': r[2] or 0,
            'fields': r[3] or 0, 'upcoming': r[4] or 0}


@conferences_bp.route('/konferensiyalar')
def conferences_hub():
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            _ensure_schema(cur)
            local = _scope_stats(cur, 'local')
            intl = _scope_stats(cur, 'international')
        conn.commit()
    finally:
        conn.close()
    return render_template('conferences_hub.html', local=local, intl=intl)


def _directory(scope):
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            _ensure_schema(cur)
            stats = _scope_stats(cur, scope)
        conn.commit()
    finally:
        conn.close()
    return render_template('conferences.html', scope=scope, stats=stats,
                           months=MONTHS_UZ, year=date.today().year,
                           is_authenticated=getattr(current_user,
                                                    'is_authenticated', False))


@conferences_bp.route('/konferensiyalar/mahalliy')
def conferences_local():
    return _directory('local')


@conferences_bp.route('/konferensiyalar/xalqaro')
def conferences_intl():
    return _directory('international')


@conferences_bp.route('/konferensiya/<slug>')
def conference_detail(slug):
    uid = current_user.id if getattr(current_user, 'is_authenticated', False) else None
    conn = get_connection()
    try:
        cur = conn.cursor(cursor_factory=psycopg2_extras.RealDictCursor)
        _ensure_schema(cur)
        cur.execute(f"SELECT {', '.join(_CONF_COLS)} FROM conferences "
                    "WHERE title_slug = %s AND is_active = TRUE", (slug,))
        row = cur.fetchone()
        if not row:
            conn.commit()
            abort(404)
        conf = _card(row)

        # o'xshashlar: bir xil scope+soha, sana bo'yicha eng yaqinlar
        cur.execute(
            f"""SELECT {', '.join(_CONF_COLS)} FROM conferences
                WHERE is_active = TRUE AND scope = %s AND id <> %s
                  AND (field = %s OR %s = '')
                ORDER BY ABS(COALESCE(start_date, CURRENT_DATE) -
                             COALESCE(%s::date, CURRENT_DATE)) ASC
                LIMIT %s""",
            (conf['scope'], conf['id'], conf['field'], conf['field'],
             conf['start_date'], SIMILAR_LIMIT))
        similar = [_card(r) for r in cur.fetchall()]

        saved = False
        has_plan = False
        in_roadmap = False
        if uid:
            cur.execute("SELECT 1 FROM user_conference_bookmarks "
                        "WHERE user_id = %s AND conference_id = %s", (uid, conf['id']))
            saved = cur.fetchone() is not None
            try:
                cur.execute("SELECT id FROM roadmap_plans "
                            "WHERE user_id = %s AND is_active LIMIT 1", (uid,))
                plan = cur.fetchone()
                has_plan = plan is not None
                if plan:
                    cur.execute("SELECT 1 FROM roadmap_conferences "
                                "WHERE plan_id = %s AND source_conference_id = %s",
                                (plan['id'], conf['id']))
                    in_roadmap = cur.fetchone() is not None
            except Exception:
                conn.rollback()  # roadmap jadvallari hali yo'q muhit
        conn.commit()
    finally:
        conn.close()

    return render_template('conference_detail.html', c=conf, similar=similar,
                           saved=saved, has_plan=has_plan, in_roadmap=in_roadmap,
                           is_authenticated=uid is not None,
                           jsonld=_conference_jsonld(conf),
                           meta_description=(
                               f"{conf['title']} — {conf['start_uz']}"
                               f"{', ' + (conf['city'] or conf['country']) if (conf['city'] or conf['country']) else ''}."
                               " Muddatlar, manba va Roadmap integratsiyasi — Olimlar.uz"))


# ── kalendar (.ics) + JSON-LD + bosh sahifa helper ──────────────────────────

SITE_URL = 'https://olimlar.uz'


def _conference_jsonld(conf):
    """schema.org Event JSON-LD (detail sahifa SEO)."""
    import json as _json
    loc_parts = [x for x in (conf.get('venue'), conf.get('city'),
                             conf.get('country')) if x]
    data = {
        '@context': 'https://schema.org', '@type': 'Event',
        'name': conf['title'],
        'url': f"{SITE_URL}/konferensiya/{conf['slug']}",
        'eventAttendanceMode': {
            'online': 'https://schema.org/OnlineEventAttendanceMode',
            'hybrid': 'https://schema.org/MixedEventAttendanceMode',
        }.get(conf.get('format'), 'https://schema.org/OfflineEventAttendanceMode'),
        'eventStatus': 'https://schema.org/EventScheduled',
    }
    if conf.get('start_date'):
        data['startDate'] = conf['start_date']
    if conf.get('end_date'):
        data['endDate'] = conf['end_date']
    if loc_parts:
        data['location'] = {'@type': 'Place', 'name': ', '.join(loc_parts),
                            'address': ', '.join(loc_parts)}
    if conf.get('organizer'):
        data['organizer'] = {'@type': 'Organization', 'name': conf['organizer']}
    if conf.get('description'):
        data['description'] = conf['description'][:500]
    return _json.dumps(data, ensure_ascii=False)


def _ics_dt(d):
    """date → ICS DATE qiymati (YYYYMMDD, kun bo'yi hodisa)."""
    return d.strftime('%Y%m%d') if d else None


def _ics_escape(s):
    return (str(s or '').replace('\\', '\\\\').replace(';', '\\;')
            .replace(',', '\\,').replace('\n', '\\n'))


@conferences_bp.route('/konferensiya/<slug>.ics')
def conference_ics(slug):
    """Kalendar fayli — VEVENT + submission_deadline uchun 7 kun oldin VALARM."""
    conn = get_connection()
    try:
        cur = conn.cursor(cursor_factory=psycopg2_extras.RealDictCursor)
        _ensure_schema(cur)
        cur.execute(f"SELECT {', '.join(_CONF_COLS)} FROM conferences "
                    "WHERE title_slug = %s AND is_active = TRUE", (slug,))
        row = cur.fetchone()
        conn.commit()
    finally:
        conn.close()
    if not row:
        abort(404)
    c = _card(row)
    start = _parse_date(c['start_date']) or date.today()
    end = _parse_date(c['end_date']) or start
    from datetime import timedelta
    dtend = end + timedelta(days=1)          # ICS DTEND — eksklyuziv
    now = datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')
    loc = ', '.join(x for x in (c.get('venue'), c.get('city'), c.get('country')) if x)
    lines = [
        'BEGIN:VCALENDAR', 'VERSION:2.0', 'PRODID:-//Olimlar.uz//Konferensiyalar//UZ',
        'CALSCALE:GREGORIAN', 'METHOD:PUBLISH', 'BEGIN:VEVENT',
        f'UID:conf-{c["id"]}@olimlar.uz', f'DTSTAMP:{now}',
        f'DTSTART;VALUE=DATE:{_ics_dt(start)}',
        f'DTEND;VALUE=DATE:{_ics_dt(dtend)}',
        f'SUMMARY:{_ics_escape(c["title"])}',
        f'URL:{SITE_URL}/konferensiya/{c["slug"]}',
    ]
    if loc:
        lines.append(f'LOCATION:{_ics_escape(loc)}')
    desc = c.get('organizer') or ''
    if c.get('submission_deadline'):
        desc += f" | Tezis muddati: {c.get('submission_uz')}"
    if desc:
        lines.append(f'DESCRIPTION:{_ics_escape(desc)}')
    # tezis muddati uchun 7 kun oldin ogohlantirish
    sub = _parse_date(c['submission_deadline'])
    if sub:
        lines += ['BEGIN:VALARM', 'ACTION:DISPLAY',
                  f'DESCRIPTION:{_ics_escape(c["title"])} — tezis muddati yaqin',
                  'TRIGGER;VALUE=DATE-TIME:'
                  + (datetime(sub.year, sub.month, sub.day)
                     - timedelta(days=7)).strftime('%Y%m%dT090000Z'),
                  'END:VALARM']
    lines += ['END:VEVENT', 'END:VCALENDAR']
    from flask import Response
    ics = '\r\n'.join(lines) + '\r\n'
    return Response(ics, mimetype='text/calendar',
                    headers={'Content-Disposition':
                             f'attachment; filename="{c["slug"]}.ics"'})


def get_upcoming_conferences(limit=3):
    """Bosh sahifa bloki uchun eng yaqin kelayotgan konferensiyalar (kesh yo'q —
    yengil so'rov). Xatoda bo'sh ro'yxat."""
    try:
        conn = get_connection()
        try:
            cur = conn.cursor(cursor_factory=psycopg2_extras.RealDictCursor)
            _ensure_schema(cur)
            cur.execute(f"""
                SELECT {', '.join(_CONF_COLS)} FROM conferences
                WHERE is_active = TRUE AND start_date IS NOT NULL
                  AND start_date >= CURRENT_DATE
                ORDER BY start_date ASC, id DESC LIMIT %s""", (limit,))
            rows = [_card(r) for r in cur.fetchall()]
            conn.commit()
            return rows
        finally:
            conn.close()
    except Exception:
        return []


# ── bookmark (user_bookmarks UX aksi) ────────────────────────────────────────

@conferences_bp.route('/api/v1/conferences/bookmark', methods=['POST'])
@csrf.exempt
@login_required
def conference_bookmark():
    data = request.get_json(silent=True) or {}
    try:
        cid = int(data.get('conference_id'))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': "Noto'g'ri so'rov"}), 400
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            _ensure_schema(cur)
            cur.execute("DELETE FROM user_conference_bookmarks "
                        "WHERE user_id = %s AND conference_id = %s",
                        (current_user.id, cid))
            if cur.rowcount:
                saved = False
            else:
                cur.execute("""
                    INSERT INTO user_conference_bookmarks (user_id, conference_id)
                    VALUES (%s, %s) ON CONFLICT DO NOTHING
                """, (current_user.id, cid))
                saved = True
            cur.execute("SELECT COUNT(*) FROM user_conference_bookmarks "
                        "WHERE user_id = %s", (current_user.id,))
            total = cur.fetchone()[0] or 0
        conn.commit()
        return jsonify({'success': True, 'saved': saved, 'total_saved': total})
    finally:
        conn.close()


# ── Roadmap integratsiyasi ───────────────────────────────────────────────────

@conferences_bp.route('/api/v1/conferences/roadmap', methods=['POST'])
@csrf.exempt
@login_required
def conference_to_roadmap():
    """Katalogdagi konferensiyani foydalanuvchining faol Roadmap rejasiga
    qo'shadi (roadmap_conferences, status='reja'). Faol reja yo'q bo'lsa
    need_plan=True — frontend /reja wizard'ga yo'naltiradi."""
    data = request.get_json(silent=True) or {}
    try:
        cid = int(data.get('conference_id'))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': "Noto'g'ri so'rov"}), 400
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            _ensure_schema(cur)
            cur.execute("SELECT id, title, city, country, start_date, "
                        "submission_deadline, title_slug FROM conferences "
                        "WHERE id = %s AND is_active = TRUE", (cid,))
            conf = cur.fetchone()
            if not conf:
                return jsonify({'success': False,
                                'error': 'Konferensiya topilmadi'}), 404
            try:
                cur.execute("SELECT id FROM roadmap_plans "
                            "WHERE user_id = %s AND is_active LIMIT 1",
                            (current_user.id,))
                plan = cur.fetchone()
            except Exception:
                plan = None
            if not plan:
                conn.rollback()
                return jsonify({'success': False, 'need_plan': True,
                                'message': 'Avval Roadmap yarating'})
            # takror qo'shishni yumshoq bloklash
            cur.execute("SELECT 1 FROM roadmap_conferences "
                        "WHERE plan_id = %s AND source_conference_id = %s",
                        (plan[0], cid))
            if cur.fetchone():
                return jsonify({'success': True, 'already': True,
                                'message': "Roadmap'ingizda allaqachon bor"})
            location = ', '.join(x for x in (conf[2], conf[3]) if x) or None
            cur.execute("""
                INSERT INTO roadmap_conferences
                    (plan_id, name, location, event_date, deadline, url,
                     source_conference_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id
            """, (plan[0], conf[1][:500], location and location[:300],
                  conf[4], conf[5], f"/konferensiya/{conf[6]}", cid))
            new_id = cur.fetchone()[0]
        conn.commit()
        return jsonify({'success': True, 'id': new_id,
                        'message': "Roadmap'ga qo'shildi"})
    finally:
        conn.close()


# ── ixtisoslik obunasi — yangi konf. alertlari ──────────────────────────────

def _match_tokens(label):
    """Obuna nomidan mazmunli tokenlar (≥6 harf, stop-so'zsiz)."""
    return [t for t in (m.group(0).lower()
                        for m in _TOKEN_RE.finditer(label or ''))
            if t not in _MATCH_STOP][:6]


def notify_conference_subscribers(conn, conf_rows):
    """Yangi konferensiyalarni mos ixtisoslik obunachilariga tarqatadi.

    Moslik best-effort: obunaning ixtisoslik_nomi tokenlaridan biri konf.
    title+field matnida uchrasa (spec: mo'rt kod-mapping majburlanmaydi).
    Dedup — conference_notifications_log. Xabar soni qaytadi; ko'tarilmaydi."""
    sent = 0
    try:
        with conn.cursor() as cur:
            _ensure_schema(cur)
            from blueprints.subscriptions import (_ensure_schema as _ensure_subs,
                                                  _telegram_chat_id)
            from blueprints.notifications import _ensure_schema as _ensure_notif
            _ensure_subs(cur)
            _ensure_notif(cur)
            cur.execute("""
                SELECT s.id, s.user_id, s.ixtisoslik, s.ixtisoslik_nomi,
                       s.notify_site, s.notify_telegram, u.telegram_chat_id, u.email
                FROM specialty_subscriptions s
                JOIN users u ON u.id = s.user_id
                WHERE s.notify_site OR s.notify_telegram
            """)
            subs = cur.fetchall()
            if not subs:
                return 0
            from blueprints.reminders import _send_telegram
            for conf in conf_rows:
                hay = f"{conf.get('title', '')} {conf.get('field', '')}".lower()
                when = conf.get('start_uz') or conf.get('start_date') or ''
                for (sub_id, user_id, code, label, n_site, n_tg,
                     chat_id, email) in subs:
                    tokens = _match_tokens(label or code)
                    if not tokens or not any(t in hay for t in tokens):
                        continue
                    cur.execute("""
                        SELECT sent_via FROM conference_notifications_log
                        WHERE subscription_id = %s AND conference_id = %s
                    """, (sub_id, conf['id']))
                    already = {r[0] for r in cur.fetchall()}
                    msg = (f"Siz obuna bo'lgan yo'nalishda yangi konferensiya: "
                           f"{conf['title']}" + (f" — {when}" if when else '') +
                           f"\n🔗 /konferensiya/{conf.get('slug', '')}")
                    if n_site and 'site' not in already:
                        cur.execute("""
                            INSERT INTO user_alerts (user_id, title, message, level)
                            VALUES (%s, %s, %s, 'info')
                        """, (user_id, '🔔 Yangi konferensiya', msg))
                        cur.execute("""
                            INSERT INTO conference_notifications_log
                                (subscription_id, conference_id, sent_via)
                            VALUES (%s, %s, 'site')
                        """, (sub_id, conf['id']))
                        sent += 1
                    tg = _telegram_chat_id(chat_id, email)
                    if n_tg and tg and 'telegram' not in already:
                        if _send_telegram(tg, {
                                'title': 'Yangi konferensiya',
                                'description': f"{conf['title']} — {when}",
                                'url': ('https://olimlar.uz/konferensiya/'
                                        + (conf.get('slug') or ''))}):
                            cur.execute("""
                                INSERT INTO conference_notifications_log
                                    (subscription_id, conference_id, sent_via)
                                VALUES (%s, %s, 'telegram')
                            """, (sub_id, conf['id']))
                            sent += 1
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
    return sent


def notify_conference_deadlines(conn):
    """Tezis muddati 14/7/1 kun qolgan konferensiyalarni bookmark qilgan
    foydalanuvchilarga eslatadi (sayt + Telegram). conference_deadline_log orqali
    dedup (har user × konf × kun-oynasi bir marta). Xabar soni qaytadi."""
    sent = 0
    try:
        with conn.cursor() as cur:
            _ensure_schema(cur)
            try:
                from blueprints.notifications import _ensure_schema as _ensure_notif
                _ensure_notif(cur)
            except Exception:
                pass
            # bugun bucketlardan biriga to'g'ri keladigan muddatli, bookmark qilingan konf.
            cur.execute("""
                SELECT b.user_id, c.id, c.title, c.title_slug, c.submission_deadline,
                       (c.submission_deadline - CURRENT_DATE) AS days,
                       u.telegram_chat_id, u.email
                FROM user_conference_bookmarks b
                JOIN conferences c ON c.id = b.conference_id
                JOIN users u ON u.id = b.user_id
                WHERE c.is_active = TRUE AND c.submission_deadline IS NOT NULL
                  AND (c.submission_deadline - CURRENT_DATE) = ANY(%s)
            """, (list(DEADLINE_BUCKETS),))
            rows = cur.fetchall()
            if not rows:
                conn.commit()
                return 0
            try:
                from blueprints.subscriptions import _telegram_chat_id
                from blueprints.reminders import _send_telegram
            except Exception:
                _telegram_chat_id = _send_telegram = None
            for user_id, cid, title, slug, deadline, days, chat_id, email in rows:
                days = int(days)
                # dedup: shu user × konf × kun-oynasi
                cur.execute("SELECT 1 FROM conference_deadline_log WHERE user_id = %s "
                            "AND conference_id = %s AND days_bucket = %s",
                            (user_id, cid, days))
                if cur.fetchone():
                    continue
                when = 'bugun' if days == 0 else f'{days} kun qoldi'
                msg = (f"Saqlab qo'ygan konferensiyangiz tezis muddati {when}: "
                       f"{title}\n🔗 /konferensiya/{slug}")
                cur.execute("INSERT INTO user_alerts (user_id, title, message, level) "
                            "VALUES (%s, %s, %s, 'warning')",
                            (user_id, '⏰ Tezis muddati yaqin', msg))
                cur.execute("INSERT INTO conference_deadline_log "
                            "(user_id, conference_id, days_bucket) VALUES (%s, %s, %s) "
                            "ON CONFLICT DO NOTHING", (user_id, cid, days))
                sent += 1
                tg = _telegram_chat_id(chat_id, email) if _telegram_chat_id else None
                if tg and _send_telegram:
                    _send_telegram(tg, {
                        'title': f'⏰ Tezis muddati {when}',
                        'description': title,
                        'url': f"{SITE_URL}/konferensiya/{slug}"})
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
    return sent


@conferences_bp.route('/api/v1/conferences/dispatch-alerts', methods=['POST'])
@csrf.exempt
def dispatch_alerts():
    """Scraper tugagach / kunlik cron GitHub Actions chaqiradi (REMINDERS_API_KEY
    bilan). Oxirgi 8 kunda qo'shilgan konferensiyalarni obunachilarga tarqatadi
    + tezis muddati 14/7/1 kun qolganlarini bookmark egalariga eslatadi. Dedup
    loglar tufayli idempotent."""
    key = request.headers.get('X-Api-Key') or request.args.get('key') or ''
    expected = os.environ.get('REMINDERS_API_KEY', '')
    if not expected or key != expected:
        return jsonify({'ok': False, 'error': 'forbidden'}), 403
    conn = get_connection()
    try:
        cur = conn.cursor(cursor_factory=psycopg2_extras.RealDictCursor)
        _ensure_schema(cur)
        cur.execute(f"""
            SELECT {', '.join(_CONF_COLS)} FROM conferences
            WHERE is_active = TRUE AND created_at >= NOW() - INTERVAL '8 days'
            ORDER BY id
        """)
        rows = [_card(r) for r in cur.fetchall()]
        conn.commit()
        sent = notify_conference_subscribers(conn, rows)
        deadline_sent = notify_conference_deadlines(conn)
        return jsonify({'ok': True, 'new_conferences': len(rows), 'sent': sent,
                        'deadline_reminders': deadline_sent})
    finally:
        conn.close()


# ── admin CRUD (admin/grants naqshi) ─────────────────────────────────────────

_ADMIN_FIELDS = ('title', 'scope', 'organizer', 'field', 'region', 'city',
                 'event_type', 'start_date', 'end_date', 'format', 'publisher',
                 'submission_deadline', 'country', 'source_url', 'description',
                 # v2 tahrirlanadigan maydonlar
                 'acronym', 'title_en', 'venue', 'organizer_contact', 'cfp_url',
                 'registration_deadline', 'notification_date', 'poster_image',
                 'ccf_rank', 'core_rank')

# _ADMIN_FIELDS'dan tashqari, alohida ishlov beriladigan ustunlar (bool/array/fk)
_ADMIN_EXTRA = ('is_scopus_indexed', 'is_multiday', 'field_id', 'tags',
                'doi_available')


def _get_conference_fields():
    """conference_fields ro'yxati (admin forma select'i uchun)."""
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            _ensure_schema(cur)
            cur.execute("SELECT id, name FROM conference_fields ORDER BY name")
            rows = [{'id': r[0], 'name': r[1]} for r in cur.fetchall()]
        conn.commit()
        return rows
    except Exception:
        return []
    finally:
        conn.close()


def _admin_form_values():
    f = request.form
    v = {k: (f.get(k) or '').strip() or None for k in _ADMIN_FIELDS}
    v['scope'] = v['scope'] if v['scope'] in SCOPES else 'local'
    for k in ('start_date', 'end_date', 'submission_deadline',
              'registration_deadline', 'notification_date'):
        v[k] = _parse_date(v[k])
    v['is_scopus_indexed'] = f.get('is_scopus_indexed') == 'on'
    v['doi_available'] = f.get('doi_available') == 'on'
    v['is_multiday'] = bool(v['start_date'] and v['end_date']
                            and v['end_date'] != v['start_date'])
    # field_id (soha select) — bo'sh bo'lsa None
    try:
        v['field_id'] = int(f.get('field_id')) if f.get('field_id') else None
    except (TypeError, ValueError):
        v['field_id'] = None
    # tags — vergul bilan ajratilgan
    tags = [t.strip() for t in (f.get('tags') or '').split(',') if t.strip()]
    v['tags'] = tags or None
    return v


@conferences_bp.route('/admin/conferences')
@login_required
def admin_conferences():
    from app import _require_admin
    _require_admin()
    q = (request.args.get('q') or '').strip()
    scope = request.args.get('scope') or ''
    where, params = ["TRUE"], []
    if q:
        where.append("(title ILIKE %s OR organizer ILIKE %s)")
        params += [f'%{q}%'] * 2
    if scope in SCOPES:
        where.append("scope = %s")
        params.append(scope)
    conn = get_connection()
    try:
        cur = conn.cursor(cursor_factory=psycopg2_extras.RealDictCursor)
        _ensure_schema(cur)
        cur.execute(
            f"SELECT {', '.join(_CONF_COLS)} FROM conferences "
            f"WHERE {' AND '.join(where)} "
            "ORDER BY is_active DESC, start_date DESC NULLS LAST, id DESC "
            "LIMIT 400", params)
        items = []
        for r in cur.fetchall():
            it = _card(r)
            it['is_active'] = bool(r['is_active'])
            items.append(it)
        conn.commit()
    finally:
        conn.close()
    return render_template('admin/conferences.html', items=items, q=q,
                           scope=scope)


@conferences_bp.route('/admin/conferences/add', methods=['GET', 'POST'])
@login_required
def admin_conferences_add():
    from app import _require_admin
    _require_admin()
    if request.method == 'POST':
        v = _admin_form_values()
        if not v['title']:
            flash('Sarlavha kiritilishi shart.', 'error')
            return render_template('admin/conference_form.html', item=v,
                                   edit_mode=False, formats=FORMATS, fields=_get_conference_fields(), event_types=EVENT_TYPES_LOCAL)
        conn = get_connection()
        try:
            with conn.cursor() as cur:
                _ensure_schema(cur)
                slug = make_slug(v['title'], cur)
                cols = list(_ADMIN_FIELDS) + list(_ADMIN_EXTRA) + ['title_slug']
                ph = ', '.join(['%s'] * len(cols))
                cur.execute(
                    f"INSERT INTO conferences ({', '.join(cols)}) VALUES ({ph})",
                    [v[c] for c in _ADMIN_FIELDS]
                    + [v[c] for c in _ADMIN_EXTRA] + [slug])
            conn.commit()
            flash("Konferensiya qo'shildi!", 'success')
            return redirect('/admin/conferences')
        except Exception as e:
            conn.rollback()
            flash('Xatolik: ' + str(e), 'error')
            return render_template('admin/conference_form.html', item=v,
                                   edit_mode=False, formats=FORMATS, fields=_get_conference_fields(), event_types=EVENT_TYPES_LOCAL)
        finally:
            conn.close()
    return render_template('admin/conference_form.html', item=None,
                           edit_mode=False, formats=FORMATS, fields=_get_conference_fields(), event_types=EVENT_TYPES_LOCAL)


@conferences_bp.route('/admin/conferences/edit/<int:cid>', methods=['GET', 'POST'])
@login_required
def admin_conferences_edit(cid):
    from app import _require_admin
    _require_admin()
    conn = get_connection()
    try:
        cur = conn.cursor(cursor_factory=psycopg2_extras.RealDictCursor)
        _ensure_schema(cur)
        if request.method == 'POST':
            v = _admin_form_values()
            if not v['title']:
                flash('Sarlavha kiritilishi shart.', 'error')
            else:
                upd_cols = list(_ADMIN_FIELDS) + list(_ADMIN_EXTRA)
                sets = ', '.join(f'{c} = %s' for c in upd_cols)
                cur.execute(
                    f"UPDATE conferences SET {sets}, updated_at = NOW() "
                    "WHERE id = %s",
                    [v[c] for c in upd_cols] + [cid])
                conn.commit()
                flash('Saqlandi!', 'success')
                return redirect('/admin/conferences')
        cur.execute(f"SELECT {', '.join(_CONF_COLS)} FROM conferences "
                    "WHERE id = %s", (cid,))
        row = cur.fetchone()
        conn.commit()
        if not row:
            abort(404)
    finally:
        conn.close()
    return render_template('admin/conference_form.html', item=dict(row),
                           edit_mode=True, formats=FORMATS, fields=_get_conference_fields(), event_types=EVENT_TYPES_LOCAL)


@conferences_bp.route('/admin/conferences/toggle/<int:cid>', methods=['POST'])
@login_required
def admin_conferences_toggle(cid):
    from app import _require_admin
    _require_admin()
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            _ensure_schema(cur)
            cur.execute("UPDATE conferences SET is_active = NOT is_active, "
                        "updated_at = NOW() WHERE id = %s "
                        "RETURNING is_active", (cid,))
            row = cur.fetchone()
        conn.commit()
        flash('Faollik o\'zgartirildi.' if row else 'Topilmadi.',
              'success' if row else 'error')
    finally:
        conn.close()
    return redirect('/admin/conferences')


# ── Excel import / export (vazirlik ro'yxati) ───────────────────────────────

# Excel sarlavhasi → conferences ustuni (normallashtirilgan kalit bo'yicha)
_IMPORT_HEADERS = {
    'mavzunomi': 'title', 'title': 'title',
    'asosiytashkilot': 'organizer', 'tashkilot': 'organizer',
    'otkazishsanasi': 'dates', 'sana': 'dates',
    'yonalishi': 'field', 'yonalish': 'field', 'soha': 'field',
    'anjumanshakli': 'event_type', 'shakl': 'event_type',
    'otkazishjoyi': 'location', 'joy': 'location',
}


def _norm_header(s):
    return re.sub(r'[^a-z0-9]', '', str(s or '').lower())


def _parse_conf_date_range(s):
    """'17.04.2026-18.04.2026' → (start, end); '20.11.2026' → (d, d). Xato → (None, None)."""
    s = str(s or '').strip()
    if not s:
        return None, None
    parts = re.findall(r'(\d{1,2})[.](\d{1,2})[.](\d{4})', s)
    out = []
    for dd, mm, yy in parts:
        try:
            out.append(date(int(yy), int(mm), int(dd)))
        except ValueError:
            pass
    if not out:
        # ISO ehtimoli
        d = _parse_date(s)
        return (d, d) if d else (None, None)
    start = out[0]
    end = out[1] if len(out) > 1 else out[0]
    return start, end


def _normalize_conf_field(name, field_map):
    """Yo'nalish nomini conference_fields'ga moslaydi → (canonical_name, field_id).
    Aniq moslik bo'lmasa token-overlap; topilmasa (asl_nom, None)."""
    raw = re.sub(r'\s+', ' ', str(name or '').strip()).rstrip(' .,')
    raw = re.sub(r"\s*Yo['ʻ]?nalishi\s*$", '', raw, flags=re.IGNORECASE).strip()
    if not raw:
        return None, None
    low = raw.lower()
    if low in field_map:
        return field_map[low]['name'], field_map[low]['id']

    # token-overlap (matn buzilgan/kesilgan bo'lsa: "San'at va madaniya t").
    # Tokenlar apostrofsiz solishtiriladi va 5-harfli prefiks bo'yicha mos keladi
    # (kesilgan so'zlar: madaniya ≈ madaniyat).
    def _toks(s):
        return [t for t in re.findall(r"[a-zа-яёўқғҳ]{4,}", s.replace("'", ''))]
    rtoks = _toks(low)
    best, best_ov = None, 0
    for key, fv in field_map.items():
        ftoks = _toks(key)
        ov = 0
        for rt in rtoks:
            if any(rt[:5] == ft[:5] for ft in ftoks):
                ov += 1
        if ov > best_ov:
            best, best_ov = fv, ov
    if best and best_ov >= 2:
        return best['name'], best['id']
    return raw, None


def _normalize_form(shakl):
    """'Anjuman,' → 'Anjuman'; 'Kongres' → 'Kongress'. event_type qaytaradi."""
    s = re.sub(r'[^\wʻʼ\' ]', '', str(shakl or '')).strip()
    if not s:
        return None
    low = s.lower()
    canon = {'anjuman': 'Anjuman', 'forum': 'Forum', 'kongres': 'Kongress',
             'kongress': 'Kongress', 'seminar': 'Seminar',
             'simpozium': 'Simpozium', 'konferensiya': 'Konferensiya'}
    return canon.get(low, s.capitalize())


def _split_location(loc):
    """'Samarqand shahri, Samarqand viloyati' → (city, region)."""
    parts = [p.strip() for p in str(loc or '').split(',') if p.strip()]
    if not parts:
        return None, None
    if len(parts) == 1:
        return parts[0], None
    return parts[0], parts[-1]


def _read_import_rows(file):
    """Excel → [{title, organizer, dates, field, event_type, location, _row}].
    Sarlavha qatorini avtomatik topadi ('Mavzu nomi' bo'yicha)."""
    import openpyxl
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    # sarlavha qatorini top
    header_idx, colmap = None, {}
    for i, row in enumerate(all_rows[:6]):
        m = {}
        for j, cell in enumerate(row):
            key = _IMPORT_HEADERS.get(_norm_header(cell))
            if key:
                m[key] = j
        if 'title' in m:
            header_idx, colmap = i, m
            break
    if header_idx is None:
        raise ValueError("Sarlavha qatori topilmadi ('Mavzu nomi' ustuni yo'q).")
    out = []
    for i in range(header_idx + 1, len(all_rows)):
        row = all_rows[i]
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue
        def g(key):
            j = colmap.get(key)
            return row[j] if (j is not None and j < len(row)) else None
        out.append({
            'title': str(g('title') or '').strip(),
            'organizer': str(g('organizer') or '').strip() or None,
            'dates': g('dates'),
            'field': g('field'), 'event_type': g('event_type'),
            'location': g('location'), '_row': i + 1,
        })
    return out


@conferences_bp.route('/admin/conferences/template')
@login_required
def admin_conferences_template():
    """Bo'sh import shabloni (.xlsx)."""
    from app import _require_admin
    _require_admin()
    import openpyxl
    from flask import Response
    import io
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Konferensiyalar'
    ws.append(['Mavzu nomi', 'Asosiy tashkilot', "O'tkazish sanasi",
               "Yo'nalishi", 'Anjuman shakli', "O'tkazish joyi"])
    ws.append(['Masalan: Zamonaviy fizika muammolari', 'Toshkent davlat universiteti',
               '17.04.2026-18.04.2026', 'Ijtimoiy-gumanitar fanlar', 'Anjuman',
               'Toshkent shahri, Toshkent viloyati'])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(buf.read(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition':
                             'attachment; filename="konferensiyalar_shablon.xlsx"'})


@conferences_bp.route('/admin/conferences/export')
@login_required
def admin_conferences_export():
    """Barcha konferensiyalarni Excel'ga eksport."""
    from app import _require_admin
    _require_admin()
    import openpyxl
    from flask import Response
    import io
    conn = get_connection()
    try:
        cur = conn.cursor(cursor_factory=psycopg2_extras.RealDictCursor)
        _ensure_schema(cur)
        cur.execute(f"SELECT {', '.join(_CONF_COLS)} FROM conferences "
                    "ORDER BY start_date DESC NULLS LAST, id DESC")
        rows = cur.fetchall()
        conn.commit()
    finally:
        conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Konferensiyalar'
    ws.append(['ID', 'Nomi', 'Tashkilot', 'Boshlanish', 'Tugash', 'Yo\'nalish',
               'Shakl', 'Shahar', 'Viloyat', 'Mamlakat', 'Tezis muddati',
               'Scopus', 'Manba', 'Faol'])
    for r in rows:
        ws.append([r['id'], r['title'], r.get('organizer'),
                   r['start_date'].isoformat() if r.get('start_date') else '',
                   r['end_date'].isoformat() if r.get('end_date') else '',
                   r.get('field'), r.get('event_type'), r.get('city'),
                   r.get('region'), r.get('country'),
                   r['submission_deadline'].isoformat() if r.get('submission_deadline') else '',
                   'ha' if r.get('is_scopus_indexed') else '', r.get('source_url'),
                   'ha' if r.get('is_active', True) else "yo'q"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(buf.read(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition':
                             'attachment; filename="konferensiyalar_export.xlsx"'})


@conferences_bp.route('/admin/conferences/import', methods=['GET', 'POST'])
@login_required
def admin_conferences_import():
    """Excel import: preview (dry-run, xato jadvali) → tasdiqlash → yozish."""
    from app import _require_admin
    _require_admin()
    if request.method == 'GET':
        return render_template('admin/conference_import.html', stage='upload')

    action = request.form.get('action')

    # ── 2-bosqich: tasdiqlash (preview'dan kelgan valid qatorlarni yozish) ──
    if action == 'confirm':
        import json as _json
        try:
            valid = _json.loads(request.form.get('valid_json') or '[]')
        except ValueError:
            valid = []
        if not valid:
            flash("Import qilinadigan qator yo'q.", 'error')
            return redirect('/admin/conferences/import')
        added = skipped = 0
        conn = get_connection()
        try:
            with conn.cursor() as cur:
                _ensure_schema(cur)
                for r in valid:
                    start = _parse_date(r.get('start_date'))
                    # dublikat: title + start_date
                    cur.execute("SELECT 1 FROM conferences WHERE LOWER(title) = LOWER(%s) "
                                "AND start_date IS NOT DISTINCT FROM %s",
                                (r['title'], start))
                    if cur.fetchone():
                        skipped += 1
                        continue
                    slug = make_slug(r['title'], cur)
                    end = _parse_date(r.get('end_date'))
                    cur.execute("""
                        INSERT INTO conferences
                            (title, title_slug, scope, organizer, field, field_id,
                             event_type, start_date, end_date, is_multiday, city,
                             region, country, source)
                        VALUES (%s,%s,'local',%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (r['title'], slug, r.get('organizer'), r.get('field'),
                          r.get('field_id'), r.get('event_type'), start, end,
                          bool(start and end and end != start),
                          r.get('city'), r.get('region'),
                          r.get('country') or "O'zbekiston", 'vazirlik'))
                    added += 1
            conn.commit()
        except Exception as e:
            conn.rollback()
            flash('Import xatosi: ' + str(e), 'error')
            return redirect('/admin/conferences/import')
        finally:
            conn.close()
        flash(f"Import yakunlandi: {added} ta qo'shildi, {skipped} ta dublikat "
              f"o'tkazib yuborildi.", 'success')
        return redirect('/admin/conferences')

    # ── 1-bosqich: preview (dry-run) ──
    file = request.files.get('file')
    if not file or not file.filename:
        flash('Fayl tanlanmadi.', 'error')
        return redirect('/admin/conferences/import')
    if not file.filename.lower().endswith(('.xlsx', '.xlsm')):
        flash('Faqat .xlsx fayl qabul qilinadi.', 'error')
        return redirect('/admin/conferences/import')
    try:
        raw_rows = _read_import_rows(file)
    except Exception as e:
        flash('Faylni o\'qib bo\'lmadi: ' + str(e), 'error')
        return redirect('/admin/conferences/import')

    # field_map: nom(lower) → {name, id}
    field_map = {f['name'].lower(): {'name': f['name'], 'id': f['id']}
                 for f in _get_conference_fields()}
    valid, errors = [], []
    seen = set()
    for r in raw_rows:
        errs = []
        title = r['title']
        if not title:
            errs.append('Mavzu nomi bo\'sh')
        start, end = _parse_conf_date_range(r['dates'])
        if r['dates'] and not start:
            errs.append(f"Sana formati noto'g'ri: {r['dates']}")
        fname, fid = _normalize_conf_field(r['field'], field_map)
        city, region = _split_location(r['location'])
        etype = _normalize_form(r['event_type'])
        dedup_key = (title.lower(), start.isoformat() if start else '')
        if title and dedup_key in seen:
            errs.append('Fayl ichida dublikat')
        seen.add(dedup_key)
        rec = {
            'title': title, 'organizer': r['organizer'],
            'start_date': start.isoformat() if start else None,
            'end_date': end.isoformat() if end else None,
            'field': fname, 'field_id': fid, 'event_type': etype,
            'city': city, 'region': region, 'country': "O'zbekiston",
            '_row': r['_row'],
        }
        if errs:
            errors.append({**rec, 'errors': errs})
        else:
            valid.append(rec)
    import json as _json
    return render_template('admin/conference_import.html', stage='preview',
                           valid=valid, errors=errors,
                           valid_json=_json.dumps(valid, ensure_ascii=False),
                           total=len(raw_rows))
